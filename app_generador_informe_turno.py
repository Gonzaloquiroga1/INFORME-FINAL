
import io
import re
from datetime import date, datetime, time, timedelta

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from pathlib import Path

TEMPLATE_CANDIDATES = [
    "F-GE-1483 (3).xlsx",
    "F-GE-1483.xlsx",
    "F-GE-1483_template.xlsx",
]

TEMPLATE_PATH = next(
    (f for f in TEMPLATE_CANDIDATES if Path(f).exists()),
    None
)

if TEMPLATE_PATH is None:
    raise FileNotFoundError(
        "No se encontró el archivo Excel de plantilla."
    )
OUTPUT_NAME = "Informe_general_de_turno_generado.xlsx"

# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------
def clean_text(value):
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()

def normalize_cc(value):
    txt = clean_text(value)
    if not txt:
        return ""
    # Keep digits and possible X; remove spaces/commas/dots
    txt = re.sub(r"[\s\.\-_,]+", "", txt)
    return txt

def pick_first_nonempty(*values):
    for v in values:
        if clean_text(v):
            return clean_text(v)
    return ""

def as_datetime_or_none(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None

def as_time_or_none(value):
    if not value:
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            pass
    return None

def get_personnel():
    # Build a single lookup from the workbook data
    workbook = load_workbook(TEMPLATE_PATH, data_only=False)
    records = []

    def add_record(cc, full_name, source):
        cc = normalize_cc(cc)
        full_name = clean_text(full_name)
        if cc and full_name:
            records.append({"cc": cc, "name": full_name, "source": source})

    # BD: C=cc, D=name, E=vinculo
    if "BD" in workbook.sheetnames:
        ws = workbook["BD"]
        for r in range(2, ws.max_row + 1):
            add_record(ws.cell(r, 3).value, ws.cell(r, 4).value, "BD")

    # BD1: A=cc, B=full name
    if "BD1" in workbook.sheetnames:
        ws = workbook["BD1"]
        for r in range(2, ws.max_row + 1):
            add_record(ws.cell(r, 1).value, ws.cell(r, 2).value, "BD1")

    # BD2: A=cc, C=apellidos, D=nombres
    if "BD2" in workbook.sheetnames:
        ws = workbook["BD2"]
        for r in range(2, ws.max_row + 1):
            cc = ws.cell(r, 1).value
            ap = ws.cell(r, 3).value
            nm = ws.cell(r, 4).value
            add_record(cc, f"{clean_text(ap)} {clean_text(nm)}", "BD2")

    # STAFF: E=CC, B=Nombre
    if "STAFF" in workbook.sheetnames:
        ws = workbook["STAFF"]
        for r in range(2, ws.max_row + 1):
            add_record(ws.cell(r, 5).value, ws.cell(r, 2).value, "STAFF")

    # Deduplicate by CC, keep first non-empty name
    df = pd.DataFrame(records)
    if df.empty:
        return pd.DataFrame(columns=["cc", "name", "source"])
    df = df[df["cc"] != ""].copy()
    df["name"] = df["name"].fillna("").map(clean_text)
    df = df.sort_values(["cc", "source"]).drop_duplicates(subset=["cc"], keep="first")
    return df.reset_index(drop=True)

@st.cache_data
def load_personnel_cache():
    df = get_personnel()
    lookup = {row.cc: row.name for row in df.itertuples(index=False)}
    return df, lookup

def lookup_name(cc, lookup):
    return lookup.get(normalize_cc(cc), "")

def build_year_schedule(ws, year):
    # Hoja6 has one yearly schedule that the report formulas use for date -> shift lookups.
    # We rewrite the date columns so any generated report for the selected year works.
    start = date(year, 1, 1)
    days = 366 if (date(year, 12, 31).toordinal() - date(year, 1, 1).toordinal() + 1) == 366 else 365

    for i in range(days):
        current = start + timedelta(days=i)
        excel_row = 3 + i

        # left block (E:G)
        ws.cell(excel_row, 5).value = current
        ws.cell(excel_row, 6).value = f'=TEXT(E{excel_row},"mmmm")'
        ws.cell(excel_row, 7).value = f'=TEXT(E{excel_row},"dddd")'

        # right block (N:P)
        ws.cell(excel_row, 14).value = current
        ws.cell(excel_row, 15).value = f'=TEXT(N{excel_row},"mmmm")'
        ws.cell(excel_row, 16).value = f'=TEXT(N{excel_row},"dddd")'

    # clean any leftover dates if workbook row count exceeds current year length
    for excel_row in range(3 + days, ws.max_row + 1):
        for c in (5, 6, 7, 14, 15, 16):
            ws.cell(excel_row, c).value = None

def set_text(ws, cell, value):
    ws[cell] = clean_text(value)

def set_raw(ws, cell, value):
    ws[cell] = value

def write_table_rows(ws, start_row, rows, mapping):
    """
    mapping: dict with keys as field names and values as (col_letter, transform_fn)
    rows: list[dict]
    """
    for idx, row in enumerate(rows):
        excel_row = start_row + idx
        for field, (col, fn) in mapping.items():
            val = row.get(field, "")
            if fn:
                val = fn(val)
            ws[f"{col}{excel_row}"] = val

def build_workbook(form):
    wb = load_workbook(TEMPLATE_PATH)

    # Make formulas recalculate when user opens the file
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    # Personnel lookup sheet used by formulas
    df_people, lookup = load_personnel_cache()
    if "Hoja14" not in wb.sheetnames:
        ws_h14 = wb.create_sheet("Hoja14")
    else:
        ws_h14 = wb["Hoja14"]
    ws_h14.delete_rows(1, ws_h14.max_row)
    ws_h14["A1"] = "Cedula"
    ws_h14["B1"] = "Nombre"
    for i, row in enumerate(df_people.itertuples(index=False), start=2):
        ws_h14[f"A{i}"] = row.cc
        ws_h14[f"B{i}"] = row.name

    # Update yearly dates in Hoja6 to make the lookup formulas work for the selected year
    if "Hoja6" in wb.sheetnames:
        build_year_schedule(wb["Hoja6"], form["fecha"].year)

    ws = wb["Informe general de turno"]

    # Header / general fields
    set_raw(ws, "E6", form["fecha"])
    set_raw(ws, "F7", form["seccion_salida"])
    set_raw(ws, "F42", form["fecha"])
    set_raw(ws, "F43", form["seccion"])
    set_raw(ws, "K43", form["turno"])
    set_text(ws, "L42", form["novedad_turno"])

    # Free text blocks
    set_text(ws, "B12", form["nov_tecnologica_1"])
    set_text(ws, "B15", form["nov_tecnologica_2"])
    set_text(ws, "B19", form["nov_administrativa"])
    set_text(ws, "B26", form["observaciones_incidentes"])

    # Incidents 1.3
    for i in range(3):
        r = 23 + i
        incident = form["incidentes"][i]
        set_text(ws, f"C{r}", incident.get("id_sur", ""))
        dt = as_datetime_or_none(incident.get("fecha_hora", ""))
        if dt:
            ws[f"F{r}"] = dt
        else:
            set_text(ws, f"F{r}", incident.get("fecha_hora", ""))
        set_text(ws, f"I{r}", incident.get("recurso", ""))

    # Continuidad personal (8 a la izquierda + 8 a la derecha)
    for i in range(8):
        r = 32 + i
        row = form["continuidad_izq"][i]
        cc = normalize_cc(row.get("cc", ""))
        set_text(ws, f"C{r}", cc)
        set_text(ws, f"E{r}", lookup_name(cc, lookup))
        set_text(ws, f"G{r}", row.get("hora_apoyo", ""))

        row2 = form["continuidad_der"][i]
        cc2 = normalize_cc(row2.get("cc", ""))
        set_text(ws, f"J{r}", cc2)
        set_text(ws, f"K{r}", lookup_name(cc2, lookup))
        set_text(ws, f"M{r}", row2.get("hora_apoyo", ""))

    # Staff attendance section
    staff_map = {
        47: ("supervisor_turno", "asiste_supervisor_turno"),
        48: ("supervisor_estacion", "asiste_supervisor_estacion"),
        54: ("monitoreo_1", "asiste_monitoreo_1"),
        56: ("monitoreo_2", "asiste_monitoreo_2"),
        62: ("seguimiento_1", "asiste_seguimiento_1"),
        64: ("seguimiento_2", "asiste_seguimiento_2"),
        67: ("mebog", "asiste_mebog"),
        67.1: ("idrd", "asiste_idrd"),  # handled separately
        69: ("mesa_1", "asiste_mesa_1"),
        70: ("mesa_2", "asiste_mesa_2"),
    }
    # Write left block
    for row_num, (name_field, assist_field) in staff_map.items():
        if row_num == 67.1:
            continue
        r = int(row_num)
        cc = normalize_cc(form[name_field])
        set_text(ws, f"F{r}", lookup_name(cc, lookup))
        set_text(ws, f"L{r}", form[assist_field])
    # Additional partner cells on the same row
    cc = normalize_cc(form["idrd"])
    set_text(ws, "L67", lookup_name(cc, lookup))
    # Mesa de ayuda rows 69 and 70 have two name positions: F and K
    cc = normalize_cc(form["mesa_3"])
    set_text(ws, "K69", lookup_name(cc, lookup))
    cc = normalize_cc(form["mesa_4"])
    set_text(ws, "K70", lookup_name(cc, lookup))

    # Late arrivals / absences / early departures
    for i in range(25):
        r = 91 + i
        late = form["llegadas"][i]
        cc = normalize_cc(late.get("cc", ""))
        set_text(ws, f"C{r}", cc)
        set_text(ws, f"D{r}", lookup_name(cc, lookup))
        set_text(ws, f"E{r}", late.get("registro_c4", ""))
        set_text(ws, f"F{r}", late.get("informa", ""))

        absn = form["ausencias"][i]
        cc2 = normalize_cc(absn.get("cc", ""))
        set_text(ws, f"H{r}", cc2)
        set_text(ws, f"J{r}", lookup_name(cc2, lookup))
        set_text(ws, f"K{r}", absn.get("tipo_evento", ""))
        set_text(ws, f"L{r}", absn.get("informa", ""))
        set_text(ws, f"M{r}", absn.get("soportes", ""))

    for i in range(10):
        r = 117 + i
        ret = form["retiros"][i]
        cc = normalize_cc(ret.get("cc", ""))
        set_text(ws, f"G{r}", cc)
        set_text(ws, f"J{r}", lookup_name(cc, lookup))
        set_text(ws, f"K{r}", ret.get("motivo", ""))
        set_text(ws, f"M{r}", ret.get("hora_salida", ""))

    # Plant support / contractor support
    for i in range(15):
        r = 130 + i
        sup = form["apoyo_planta"][i]
        cc = normalize_cc(sup.get("cc", ""))
        set_text(ws, f"C{r}", cc)
        set_text(ws, f"D{r}", lookup_name(cc, lookup))
        set_text(ws, f"E{r}", sup.get("registro_c4", ""))
        set_text(ws, f"F{r}", sup.get("motivo", ""))
        set_text(ws, f"H{r}", sup.get("observaciones", ""))
        set_text(ws, f"M{r}", sup.get("rol", ""))

        r2 = 191 + i
        sup2 = form["apoyo_contratista"][i]
        cc2 = normalize_cc(sup2.get("cc", ""))
        set_text(ws, f"C{r2}", cc2)
        set_text(ws, f"D{r2}", lookup_name(cc2, lookup))
        set_text(ws, f"E{r2}", sup2.get("registro_c4", ""))
        set_text(ws, f"F{r2}", sup2.get("motivo", ""))
        set_text(ws, f"H{r2}", sup2.get("observaciones", ""))
        set_text(ws, f"M{r2}", sup2.get("rol", ""))

    # Convenience: set the date style on some cells
    for cell in ("E6", "F42"):
        try:
            ws[cell].number_format = "dd/mm/yyyy"
        except Exception:
            pass

    # Also make the day-name cells easier to read when workbook is opened
    for cell in ("F6", "H42"):
        try:
            ws[cell].number_format = "[$-es-CO]dddd"
        except Exception:
            pass

    # Keep alignment nice for wrapped text blocks
    for cell in ("B12", "B15", "B19", "B26"):
        ws[cell].alignment = Alignment(wrap_text=True, vertical="top")

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def empty_row():
    return {"cc": "", "hora_apoyo": ""}

def empty_late_row():
    return {"cc": "", "registro_c4": "", "informa": ""}

def empty_abs_row():
    return {"cc": "", "tipo_evento": "", "informa": "", "soportes": ""}

def empty_ret_row():
    return {"cc": "", "motivo": "", "hora_salida": ""}

def empty_sup_row():
    return {"cc": "", "registro_c4": "", "motivo": "", "observaciones": "", "rol": ""}

def default_form():
    today = date.today()
    return {
        "fecha": today,
        "seccion_salida": 1,
        "seccion": 1,
        "turno": 1,
        "novedad_turno": "SIN NOVEDAD",
        "nov_tecnologica_1": "",
        "nov_tecnologica_2": "",
        "nov_administrativa": "",
        "observaciones_incidentes": "",
        "incidentes": [
            {"id_sur": "", "fecha_hora": "", "recurso": ""},
            {"id_sur": "", "fecha_hora": "", "recurso": ""},
            {"id_sur": "", "fecha_hora": "", "recurso": ""},
        ],
        "continuidad_izq": [empty_row() for _ in range(8)],
        "continuidad_der": [empty_row() for _ in range(8)],
        "supervisor_turno": "",
        "asiste_supervisor_turno": "",
        "supervisor_estacion": "",
        "asiste_supervisor_estacion": "",
        "monitoreo_1": "",
        "asiste_monitoreo_1": "",
        "monitoreo_2": "",
        "asiste_monitoreo_2": "",
        "seguimiento_1": "",
        "asiste_seguimiento_1": "",
        "seguimiento_2": "",
        "asiste_seguimiento_2": "",
        "mebog": "",
        "asiste_mebog": "",
        "idrd": "",
        "asiste_idrd": "",
        "mesa_1": "",
        "asiste_mesa_1": "",
        "mesa_2": "",
        "asiste_mesa_2": "",
        "mesa_3": "",
        "mesa_4": "",
        "llegadas": [empty_late_row() for _ in range(25)],
        "ausencias": [empty_abs_row() for _ in range(25)],
        "retiros": [empty_ret_row() for _ in range(10)],
        "apoyo_planta": [empty_sup_row() for _ in range(15)],
        "apoyo_contratista": [empty_sup_row() for _ in range(15)],
    }

def ensure_state():
    if "form" not in st.session_state:
        st.session_state.form = default_form()

def build_df_from_rows(rows, columns):
    return pd.DataFrame([{c: r.get(c, "") for c in columns} for r in rows])

def update_rows_from_df(df, rows, columns):
    out = []
    for _, row in df.iterrows():
        item = {c: row.get(c, "") for c in columns}
        out.append(item)
    # pad
    while len(out) < len(rows):
        out.append({c: "" for c in columns})
    return out[: len(rows)]

# ---------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------
st.set_page_config(page_title="Generador de Informe de Turno", layout="wide")
ensure_state()
people_df, people_lookup = load_personnel_cache()

st.title("Generador del informe general de turno")

st.caption(
    "Llena los datos principales y exporta una copia diligenciada del archivo base."
)

with st.sidebar:
    st.subheader("Archivo base")
    st.write("Plantilla cargada: `F-GE-1483_template.xlsx`")
    st.write(f"Registros de personal encontrados: {len(people_df):,}".replace(",", "."))
    st.download_button(
        "Descargar plantilla base",
        data=open(TEMPLATE_PATH, "rb").read(),
        file_name=TEMPLATE_PATH,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

form = st.session_state.form

col1, col2, col3 = st.columns(3)
with col1:
    form["fecha"] = st.date_input("Fecha del informe", value=form["fecha"])
with col2:
    form["seccion_salida"] = st.selectbox("Sección que sale", [1, 2, 3, 4, 5], index=form["seccion_salida"] - 1)
with col3:
    form["seccion"] = st.selectbox("Sección del turno", [1, 2, 3, 4, 5], index=form["seccion"] - 1)

form["turno"] = st.selectbox("Turno", [1, 2, 3, 4, 5], index=form["turno"] - 1)
form["novedad_turno"] = st.text_input("Novedad en turno", value=form["novedad_turno"])

tab1, tab2, tab3, tab4 = st.tabs(["Novedades", "Personal", "Novedades de personal", "Apoyos"])

with tab1:
    c1, c2 = st.columns(2)
    with c1:
        form["nov_tecnologica_1"] = st.text_area(
            "1.1 Novedades tecnológicas - bloque 1",
            value=form["nov_tecnologica_1"],
            height=120,
        )
        form["nov_tecnologica_2"] = st.text_area(
            "1.1 Novedades tecnológicas - bloque 2",
            value=form["nov_tecnologica_2"],
            height=120,
        )
    with c2:
        form["nov_administrativa"] = st.text_area(
            "1.2 Novedades administrativas",
            value=form["nov_administrativa"],
            height=260,
        )
        form["observaciones_incidentes"] = st.text_area(
            "Observaciones de incidentes / novedades del personal",
            value=form["observaciones_incidentes"],
            height=120,
        )

    st.subheader("1.3 Incidentes de alto impacto")
    inc_df = pd.DataFrame(form["incidentes"])
    inc_df = st.data_editor(
        inc_df,
        num_rows="fixed",
        use_container_width=True,
        column_config={
            "id_sur": st.column_config.TextColumn("ID SUR"),
            "fecha_hora": st.column_config.TextColumn("Fecha - hora"),
            "recurso": st.column_config.TextColumn("Agencias desplegadas y recurso asignado"),
        },
        key="incidentes_editor",
    )
    form["incidentes"] = inc_df.to_dict(orient="records")

with tab2:
    st.write("Ingresa las cédulas; el sistema completa el nombre al exportar.")
    c1, c2 = st.columns(2)

    with c1:
        st.markdown("**Continuidad – bloque izquierdo**")
        cont_left = pd.DataFrame(form["continuidad_izq"])
        cont_left = st.data_editor(
            cont_left,
            num_rows="fixed",
            use_container_width=True,
            column_config={
                "cc": st.column_config.TextColumn("Cédula"),
                "hora_apoyo": st.column_config.TextColumn("Hora final de apoyo"),
            },
            key="cont_left",
        )
        form["continuidad_izq"] = cont_left.to_dict(orient="records")

    with c2:
        st.markdown("**Continuidad – bloque derecho**")
        cont_right = pd.DataFrame(form["continuidad_der"])
        cont_right = st.data_editor(
            cont_right,
            num_rows="fixed",
            use_container_width=True,
            column_config={
                "cc": st.column_config.TextColumn("Cédula"),
                "hora_apoyo": st.column_config.TextColumn("Hora final de apoyo"),
            },
            key="cont_right",
        )
        form["continuidad_der"] = cont_right.to_dict(orient="records")

    st.markdown("**Personal de turno / asistencia**")
    staff_fields = [
        ("supervisor_turno", "Asistente supervisor turno"),
        ("asiste_supervisor_turno", "¿Asiste?"),
        ("supervisor_estacion", "Supervisor estación"),
        ("asiste_supervisor_estacion", "¿Asiste?"),
        ("monitoreo_1", "Monitoreo 1"),
        ("asiste_monitoreo_1", "¿Asiste?"),
        ("monitoreo_2", "Monitoreo 2"),
        ("asiste_monitoreo_2", "¿Asiste?"),
        ("seguimiento_1", "Seguimiento 1"),
        ("asiste_seguimiento_1", "¿Asiste?"),
        ("seguimiento_2", "Seguimiento 2"),
        ("asiste_seguimiento_2", "¿Asiste?"),
        ("mebog", "MEBOG"),
        ("asiste_mebog", "¿Asiste?"),
        ("idrd", "IDRD"),
        ("asiste_idrd", "¿Asiste?"),
        ("mesa_1", "Mesa de ayuda 1"),
        ("asiste_mesa_1", "¿Asiste?"),
        ("mesa_2", "Mesa de ayuda 2"),
        ("asiste_mesa_2", "¿Asiste?"),
        ("mesa_3", "Mesa de ayuda 3"),
        ("mesa_4", "Mesa de ayuda 4"),
    ]
    cols = st.columns(2)
    for idx, (field, label) in enumerate(staff_fields):
        with cols[idx % 2]:
            form[field] = st.text_input(label, value=form[field])

with tab3:
    st.subheader("5.1 Llegadas tarde")
    late_df = pd.DataFrame(form["llegadas"])
    late_df = st.data_editor(
        late_df,
        num_rows="fixed",
        use_container_width=True,
        column_config={
            "cc": st.column_config.TextColumn("Cédula"),
            "registro_c4": st.column_config.TextColumn("Registro ingreso C4"),
            "informa": st.column_config.TextColumn("Informa"),
        },
        key="late_editor",
    )
    form["llegadas"] = late_df.to_dict(orient="records")

    st.subheader("5.2 Ausentismos")
    abs_df = pd.DataFrame(form["ausencias"])
    abs_df = st.data_editor(
        abs_df,
        num_rows="fixed",
        use_container_width=True,
        column_config={
            "cc": st.column_config.TextColumn("Cédula"),
            "tipo_evento": st.column_config.TextColumn("Tipo de evento reportado"),
            "informa": st.column_config.TextColumn("Informa"),
            "soportes": st.column_config.TextColumn("Soportes"),
        },
        key="abs_editor",
    )
    form["ausencias"] = abs_df.to_dict(orient="records")

    st.subheader("5.3 Retiro anticipado")
    ret_df = pd.DataFrame(form["retiros"])
    ret_df = st.data_editor(
        ret_df,
        num_rows="fixed",
        use_container_width=True,
        column_config={
            "cc": st.column_config.TextColumn("Cédula"),
            "motivo": st.column_config.TextColumn("Motivo retiro"),
            "hora_salida": st.column_config.TextColumn("Hora salida"),
        },
        key="ret_editor",
    )
    form["retiros"] = ret_df.to_dict(orient="records")

with tab4:
    st.subheader("6. Personal de apoyo planta")
    sup_df = pd.DataFrame(form["apoyo_planta"])
    sup_df = st.data_editor(
        sup_df,
        num_rows="fixed",
        use_container_width=True,
        column_config={
            "cc": st.column_config.TextColumn("Cédula"),
            "registro_c4": st.column_config.TextColumn("Registro ingreso C4"),
            "motivo": st.column_config.TextColumn("Motivo del apoyo"),
            "observaciones": st.column_config.TextColumn("Observaciones"),
            "rol": st.column_config.TextColumn("Rol"),
        },
        key="sup_editor",
    )
    form["apoyo_planta"] = sup_df.to_dict(orient="records")

    st.subheader("7. Personal de apoyo contratista")
    sup2_df = pd.DataFrame(form["apoyo_contratista"])
    sup2_df = st.data_editor(
        sup2_df,
        num_rows="fixed",
        use_container_width=True,
        column_config={
            "cc": st.column_config.TextColumn("Cédula"),
            "registro_c4": st.column_config.TextColumn("Registro ingreso C4"),
            "motivo": st.column_config.TextColumn("Motivo del apoyo"),
            "observaciones": st.column_config.TextColumn("Observaciones"),
            "rol": st.column_config.TextColumn("Rol"),
        },
        key="sup2_editor",
    )
    form["apoyo_contratista"] = sup2_df.to_dict(orient="records")

st.session_state.form = form

def calculate_turn(fecha, seccion):
    # Matches the cyclic turn pattern in Hoja6.
    base = {1: 2, 2: 1, 3: 5, 4: 4, 5: 3}
    start = date(2024, 1, 1)
    offset = (fecha - start).days
    b = base.get(int(seccion), 1)
    return ((b - 1 - offset) % 5) + 1

# update derived turn automatically for preview/generation
derived_turn = calculate_turn(form["fecha"], form["seccion"])
st.info(f"Turno calculado según la plantilla: **{derived_turn}**")
form["turno"] = derived_turn

if st.button("Generar archivo Excel", type="primary"):
    output = build_workbook(form)
    st.success("Archivo generado correctamente.")
    st.download_button(
        "Descargar informe generado",
        data=output.getvalue(),
        file_name=OUTPUT_NAME,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
